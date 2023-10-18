import csv
import sys
from enum import Enum
from datetime import date
import re
from typing import List
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from datetime import datetime
from openpyxl.chart import (
    PieChart,
    Reference
)

# -- Title --
SHEET_TITLES = ['Titre', 'Type', 'Catégorie', 'Compte', 'Valeur', 'Vérification', 'Date Prélèv.', 'Date Emission', 'Description']
SHEET_TITLE_STYLE = Font(color='FFFFFF')
SHEET_TITLE_FILL = PatternFill("solid", fgColor="A5A5A5")
SHEET_TITLE_BORDER = Border(bottom=Side(border_style="medium", color="000000"))

# -- Row --
SHEET_ROW_BASIC_COLOR = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")
SHEET_ROW_GREY_COLOR = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
SHEET_ROW_RED_FONT_COLOR = Font(color="C65911")
SHEET_ROW_BLUE_FONT_COLOR = Font(color="4772C4")
SHEET_ROW_RED_COLOR = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
SHEET_ROW_BLUE_COLOR = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
SHEET_ROW_OK_COLOR = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
SHEET_ROW_NOK_COLOR = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
SHEET_ROW_START_BORDER = Border(left=Side(border_style="medium", color="000000"), bottom=Side(border_style="thin", color="000000"), right=Side(border_style="dotted", color="000000"))
SHEET_ROW_MIDDLE_BORDER = Border(left=Side(border_style="dotted", color="000000"), bottom=Side(border_style="thin", color="000000"), right=Side(border_style="dotted", color="000000"))
SHEET_ROW_END_BORDER = Border(left=Side(border_style="dotted", color="000000"), bottom=Side(border_style="thin", color="000000"), right=Side(border_style="medium", color="000000"))

# -- VARIOUS CONST --
FILE_TO_OPEN = 'Dépenses.xlsx'

START_WORKSHEET_YEAR = 2022
START_WORKSHEET_MONTH = 9
END_WORKSHEET_YEAR = 2023 # TODO Remove
END_WORKSHEET_MONTH = 8 # TODO Remove

# -- List type depense --
SHEET_SOMME_TYPE_DEPENSE = [
    ('=Data!C2', '=SUMIF($C$2:$C$150,K3,$E$2:$E$150)'),
    ('=Data!C3', '=SUMIF($C$2:$C$150,K4,$E$2:$E$150)'),
    ('=Data!C4', '=SUMIF($C$2:$C$150,K5,$E$2:$E$150)'),
    ('=Data!C5', '=SUMIF($C$2:$C$150,K6,$E$2:$E$150)'),
    ('=Data!C6', '=SUMIF($C$2:$C$150,K7,$E$2:$E$150)'),
    ('=Data!C7', '=SUMIF($C$2:$C$150,K8,$E$2:$E$150)'),
    ('=Data!C8', '=SUMIF($C$2:$C$150,K9,$E$2:$E$150)'),
    ('=Data!C9', '=SUMIF($C$2:$C$150,K10,$E$2:$E$150)')]

class TypeEnum(Enum) :
    Entre = "Entré"
    Sortie = "Sortie"
    TransfereSortie = "Transfère Sortie"
    TransfereEntre = "Transfère Entré"
    Empty = ""

class MonthEnum(Enum):
    Janvier= 1
    Février= 2
    Mars= 3
    Avril= 4
    Mai = 5
    Juin= 6
    Juillet = 7
    Aout = 8
    Septembre = 9
    Octobre = 10
    Novembre = 11
    Décembre = 12

class CompteEnum(Enum) :
    LCL = "LCL"
    Bourso = "Boursorama"

class Operation:
    def __init__(self, type : TypeEnum, description :str, montant : float, date : date, compte : CompteEnum):
        self.type = type
        self.description = description
        self.montant = montant
        self.date = date
        self.compte = compte
    def __repr__(self):
        return repr((self.type, self.description, self.montant, self.date, self.compte))

# Main function
def main():
    # Get the arguments
    print(getOptions(sys.argv))
    input, output, lcl, bourso = getOptions(sys.argv)
    
    print (input)

    # Read the operations
    operations : List[Operation] = []
    if(lcl):
        operations.extend(convertLCLFile(lcl))
    if(bourso):
        operations.extend(convertBoursoFile(bourso))

    print('Find', len(operations), 'operation.s')

    # Sort the operations
    sortedOperations = sorted(operations, key=lambda operation: operation.date)
    exportOperations(sortedOperations)
    compteExcel(sortedOperations, input, output)

# get the options of the file
def getOptions(arguments : List[str]):
    lcl : str = None
    bourso : str = None
    input: str = None
    output : str = None
    
    for pointer in range(1, len(arguments), 2):
        if arguments[pointer] == "--lcl" :
            lcl = arguments[pointer + 1]
        elif arguments[pointer] == '--bourso' :
            bourso = arguments[pointer + 1]
        elif arguments[pointer] == '-i':
            input = arguments[pointer + 1]
        elif arguments[pointer] == '-o':
            output = arguments[pointer + 1]
        else :
            print('Unknow argument : ', arguments[pointer])
            print('Usage : python compte.py [--lcl <lcl_csv_path>, --bourso <boursorama_csv_path>, -i <input_specified_file>, -o <output_specified_file>]')
            exit(-1)
    
    return (input, output, lcl, bourso)

# Convert a csv export form lcl to use it in the program
def convertLCLFile(file: str) -> List[Operation]:
    with open(file) as csvfile:
        lclLine = csv.reader(csvfile, delimiter=';')
        lclOperation : List[Operation] = []
        for id, row in enumerate(lclLine) :
            if len(row) == 8:
                day, month, year = re.search("(\d{2})\/(\d{2})\/(\d{4})",row[0]).groups()
                dateEvent = date(int(year), int(month), int(day))
                montant = abs(float(row[1].replace(',','.')))
                description = row[4]+row[5]
                type = TypeEnum.Empty if re.match("VIR", description) else TypeEnum.Sortie if re.match("-", row[1]) else TypeEnum.Entre
                lclOperation.append(Operation(type, description, str(montant).replace('.', ','), dateEvent, CompteEnum.LCL))
        return lclOperation

# Convert a csv export from boursorama to use in the program
def convertBoursoFile(file : str) -> List[Operation]:
    with open(file) as csvfile:
        boursoline = csv.reader(csvfile, delimiter=';')
        boursoOperation : List[Operation]= []
        for id, row in enumerate(boursoline):
            if id != 0 :
                year, month, day = re.search("(\d{4})-(\d{2})-(\d{2})", row[0]).groups()
                dateEvent = date(int(year), int(month), int(day))
                montant = abs(float(row[5].replace(',','.')))
                description = row[2]
                type = TypeEnum.Empty if re.match("VIR", description) else TypeEnum.Sortie if re.match("-", row[5]) else TypeEnum.Entre
                boursoOperation.append(Operation(type, description, str(montant).replace('.', ','), dateEvent, CompteEnum.Bourso))
        return boursoOperation

# Export the operations as a CSV file
def exportOperations(operations : List[Operation]) -> None:
    with open('./export.csv', 'w', newline='') as file:
        writer = csv.writer(file)
        for id, row in enumerate(operations):
            writer.writerow(["", row.type.value,"", row.compte.value, row.montant, "NOK", row.date, "", row.description])

# Update the excel file
def compteExcel(operations : List[Operation], input : str, output : str) -> None:
    # Load the excel file 
    fileToOpen : str = input if input else FILE_TO_OPEN 
    workbook = openpyxl.load_workbook(fileToOpen, keep_links=False)

    # Manage Style of old workbooks
    yearToFind : int = START_WORKSHEET_YEAR
    monthToFind : int = START_WORKSHEET_MONTH # TODO Change with 1
    firstWorkbookFind : bool = False
    while not firstWorkbookFind or findSheet(workbook, monthToFind, yearToFind) and (monthToFind != END_WORKSHEET_MONTH or yearToFind != END_WORKSHEET_YEAR): # TODO Remove END_WORKSHEET... Tests
        if firstWorkbookFind :
            manageStyle(workbook, monthToFind, yearToFind)
        else :
            firstWorkbookFind = findSheet(workbook, monthToFind, yearToFind)
        if monthToFind == 12:
            monthToFind = 1
            yearToFind += 1
        else :
            monthToFind += 1

    # Split operations by month/year
    operationsByMonthYear : dict[date, List[Operation]] = {}
    for operation in operations :
        monthDate = date(operation.date.year, operation.date.month, 1)
        if(not operationsByMonthYear.get(monthDate)) :
            operationsByMonthYear[monthDate] = []
        operationsByMonthYear[monthDate].append(operation)

    # Add the operations
    for key in operationsByMonthYear :
        keyYear = key.year
        keyMonth = key.month
        # Create sheet if not allready exist
        if (keyYear == yearToFind and keyMonth > monthToFind) or (keyYear > yearToFind):
            createNewSheet(workbook, keyMonth, keyYear)
            manageStyle(workbook, keyMonth, keyYear)
        # Add operations
        addOperations(workbook, keyMonth, keyYear, operationsByMonthYear.get(key))

    # Save and close the new excel file
    fileToSave = output if output else 'Dépenses_'+str(datetime.timestamp(datetime.now()))+".xlsx"
    workbook.save(fileToSave)
    workbook.close()

# TODO Fix the function
# Find if a sheet exist depending of this month & year
def findSheet(workbook : Workbook, month : int, year : int):
    sheetToFind =  MonthEnum._value2member_map_[month].name + ' ' + str(year)
    return workbook.sheetnames.index(sheetToFind)

# Create a new sheet of data at the expected month & year
def createNewSheet(workbook : Workbook, month: int, year : int):
    title : str = MonthEnum._value2member_map_[month].name + ' ' + str(year)
    workbook.create_sheet(title)
    sheet = workbook[title]

    # Create sheet titles
    for id, val in enumerate(SHEET_TITLES):
        sheet[chr(65+id)+'1'].font = SHEET_TITLE_STYLE
        sheet[chr(65+id)+'1'].fill = SHEET_TITLE_FILL
        sheet[chr(65+id)+'1'].border = SHEET_TITLE_BORDER
        sheet[chr(65+id)+'1'] = val
    
    # Create sheet type de dépense
    for id, (name, value) in enumerate(SHEET_SOMME_TYPE_DEPENSE):
        sheet['K'+str(id+3)] = name
        sheet['L'+str(id+3)] = value
    
    # Create camembert
    pie = PieChart()
    labels = Reference(sheet, min_col=11, min_row=3, max_row=10)
    data = Reference(sheet, min_col=12, min_row=3, max_row=10)
    pie.add_data(data)
    pie.set_categories(labels)
    sheet.add_chart(pie, "N2")

    # Add row to statistiques
    statistiqueSheet = workbook['Statistiques']
    rowToAddIndex = (year - 2023) * 13 + 7 + month
    if month == 1 :
        statistiqueSheet.merge_cells(start_row = rowToAddIndex - 1, start_column = 1, end_row=rowToAddIndex-1, end_column=4)
        statistiqueSheet.cell(row=rowToAddIndex - 1, column=1).value = year
    statistiqueSheet.cell(row=rowToAddIndex, column = 1).value = MonthEnum._value2member_map_[month].name
    statistiqueSheet.cell(row=rowToAddIndex, column = 2).value = "=SUMIF('"+title+"'!$B$2:$B$200, Data!$A$2, '"+title+"'!$E$2:$E$200)"
    statistiqueSheet.cell(row=rowToAddIndex, column = 3).value = "=SUMIF('"+title+"'!$B$2:$B$200, Data!$A$3, '"+title+"'!$E$2:$E$200)"
    statistiqueSheet.cell(row=rowToAddIndex, column = 4).value = "="+statistiqueSheet.cell(row=rowToAddIndex, column = 2).coordinate+"-"+statistiqueSheet.cell(row=rowToAddIndex, column = 3).coordinate

# Add the information of the operations
def addOperations(workbook : Workbook, month : int, year : int, operationList : List[Operation]):
    # Open the sheet to find
    sheetToFind : str =  MonthEnum._value2member_map_[month].name + ' ' + str(year)
    sheet = workbook[sheetToFind]

    # Find the line where there in no data
    rowEmpty : int = 2 # begin at line 2, line 1 have the titles
    while (sheet.cell(row=rowEmpty, column=5).value !=  None): # See if the cell "Valeur" is empty
        rowEmpty += 1
    
    # Add the data
    for operation in operationList: # Loop throw each operation
        # Create the line
        sheet.cell(row=rowEmpty, column=2).value = operation.type.value
        sheet.cell(row=rowEmpty, column=4).value = operation.compte.value
        sheet.cell(row=rowEmpty, column=5).data_type = 'n'
        sheet.cell(row=rowEmpty, column=5).value = float(operation.montant.replace(',','.'))
        sheet.cell(row=rowEmpty, column=6).value = "NOK"
        sheet.cell(row=rowEmpty, column=7).value = operation.date.strftime("%d/%m/%Y")
        sheet.cell(row=rowEmpty, column=9).value = operation.description
        # Increment the line to implement
        rowEmpty += 1

# Manage the style of a page
def manageStyle(workbook : Workbook, month : int, year : int):
    sheetToFind =  MonthEnum._value2member_map_[month].name + ' ' + str(year)
    sheet = workbook[sheetToFind]

    # Manage column width
    sheet.column_dimensions['A'].width = 32
    sheet.column_dimensions['B'].width = 16
    sheet.column_dimensions['C'].width = 16
    sheet.column_dimensions['D'].width = 11
    sheet.column_dimensions['E'].width = 11
    sheet.column_dimensions['F'].width = 11
    sheet.column_dimensions['G'].width = 11
    sheet.column_dimensions['H'].width = 11
    sheet.column_dimensions['I'].width = 48

    # Manage line style
    # Entré/Sortie to Valeur
    sheet.conditional_formatting.add("B2:E200", FormulaRule(formula=['$B2=Data!$A$5'], stopIfTrue=True, fill=SHEET_ROW_GREY_COLOR, font=SHEET_ROW_BLUE_FONT_COLOR))
    sheet.conditional_formatting.add("B2:E200", FormulaRule(formula=['$B2=Data!$A$4'], stopIfTrue=True, fill=SHEET_ROW_GREY_COLOR, font=SHEET_ROW_RED_FONT_COLOR))
    sheet.conditional_formatting.add("B2:E200", FormulaRule(formula=['$B2=Data!$A$2'], stopIfTrue=True, fill=SHEET_ROW_BLUE_COLOR))
    sheet.conditional_formatting.add("B2:E200", FormulaRule(formula=['$B2=Data!$A$3'], stopIfTrue=True, fill=SHEET_ROW_RED_COLOR))
    sheet.conditional_formatting.add("B2:E200", FormulaRule(formula=['ISBLANK($B2)'], stopIfTrue=True, fill=SHEET_ROW_GREY_COLOR))
    
    # Vérification
    sheet.conditional_formatting.add("F2:F200", CellIsRule(operator='equal', formula=["Data!$F$3"], stopIfTrue=True, fill=SHEET_ROW_NOK_COLOR))
    sheet.conditional_formatting.add("F2:F200", CellIsRule(operator='equal', formula=["Data!$F$2"], stopIfTrue=True, fill=SHEET_ROW_OK_COLOR))
    sheet.conditional_formatting.add("F2:F200", FormulaRule(formula=['ISBLANK(F2)'], stopIfTrue=True, fill=SHEET_ROW_GREY_COLOR))

    # Other columns
    for rowIndex in range(2,201) :
        # Fill
        sheet.cell(row=rowIndex, column=1).fill = SHEET_ROW_BASIC_COLOR
        sheet.cell(row=rowIndex, column=7).fill = SHEET_ROW_BASIC_COLOR
        sheet.cell(row=rowIndex, column=8).fill = SHEET_ROW_BASIC_COLOR
        sheet.cell(row=rowIndex, column=9).fill = SHEET_ROW_BASIC_COLOR

        # Border
        sheet.cell(row=rowIndex, column=1).border = SHEET_ROW_START_BORDER
        for columnIndex in range(2,9):
            sheet.cell(row=rowIndex, column=columnIndex).border = SHEET_ROW_MIDDLE_BORDER
        sheet.cell(row=rowIndex, column=9).border = SHEET_ROW_END_BORDER

    # Manage row data validation
    # Type
    ruleType = DataValidation(type="list", formula1="Data!$A$2:$A$5")
    ruleType.add("B2:B200")
    sheet.add_data_validation(ruleType)
    # Catégorie
    ruleCategorie = DataValidation(type="list", formula1="IF($B2 = Data!$A$2,Data!$B$2:$B$5,IF($B2 = Data!$A$3,Data!$C$2:$C$9,))")
    ruleCategorie.add("C2:C200")
    sheet.add_data_validation(ruleCategorie)
    # Compte
    ruleCompte = DataValidation(type="list", formula1="Data!$D$2:$D$6")
    ruleCompte.add("D2:D200")
    sheet.add_data_validation(ruleCompte)
    # Vérification
    ruleVerification = DataValidation(type="list", formula1="Data!$F$2:$F$3")
    ruleVerification.add("F2:F200")
    sheet.add_data_validation(ruleVerification)

if __name__ == "__main__":
    main()